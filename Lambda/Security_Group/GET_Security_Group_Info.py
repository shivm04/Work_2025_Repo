import boto3
import json
import logging
import base64
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger()
logger.setLevel(logging.INFO)

ec2 = boto3.client("ec2")
ses = boto3.client("ses")
sts = boto3.client("sts")

# CONFIG
SENDER = "04shivm@gmail.com"
RECIPIENTS = ["04shivm@gmail.com"]
SUBJECT = "AWS Security Group Inbound & Outbound Rules Report"
ATTACHMENT_FILENAME = "security_group_rules.xlsx"

PHI_LOGO_URL = "https://test.com/wp-content/uploads/2024/09/logo-300x170.png.webp"


# ----------------------------------------------------------------------
# EXCEL HELPERS
# ----------------------------------------------------------------------
def style_header(ws):
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFEEEEEE", fill_type="solid")
    thin = Side(border_style="thin", color="FFBBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")


def border_table(ws):
    thin = Side(border_style="thin", color="FFBBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left")


def autosize(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            text = str(cell.value or "")
            col = cell.column
            dims[col] = max(dims.get(col, 0), len(text))

    for col, width in dims.items():
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = min(max(width + 2, 8), 60)


# ----------------------------------------------------------------------
# BUILD WORKBOOK
# ----------------------------------------------------------------------
def build_workbook(inbound_rows, outbound_rows):
    wb = Workbook()
    headers = ["Sr. No", "Security Group Name", "Security Group ID", "Type",
               "Port Range", "Protocol", "Target", "Description"]

    # INBOUND SHEET
    ws_in = wb.active
    ws_in.title = "Inbound Rules"
    ws_in.append(headers)

    for r in inbound_rows:
        if r.get("Separator"):
            ws_in.append([""] * len(headers))
            row_idx = ws_in.max_row
            for col in range(1, len(headers) + 1):
                cell = ws_in.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFF0F0F0", fill_type="solid")
        else:
            ws_in.append([
                r.get("SrNo"), r.get("GroupName"), r.get("GroupId"), r.get("Type"),
                r.get("PortRange"), r.get("Protocol"), r.get("Target"), r.get("Description")
            ])

    style_header(ws_in)
    border_table(ws_in)
    autosize(ws_in)

    # OUTBOUND SHEET
    ws_out = wb.create_sheet("Outbound Rules")
    ws_out.append(headers)

    for r in outbound_rows:
        if r.get("Separator"):
            ws_out.append([""] * len(headers))
            row_idx = ws_out.max_row
            for col in range(1, len(headers) + 1):
                cell = ws_out.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFF0F0F0", fill_type="solid")
        else:
            ws_out.append([
                r.get("SrNo"), r.get("GroupName"), r.get("GroupId"), r.get("Type"),
                r.get("PortRange"), r.get("Protocol"), r.get("Target"), r.get("Description")
            ])

    style_header(ws_out)
    border_table(ws_out)
    autosize(ws_out)

    return wb


# ----------------------------------------------------------------------
# RAW EMAIL BUILDER
# ----------------------------------------------------------------------
def build_raw_email(sender, recipients, subject, html, attachment_bytes, filename):
    boundary = "NextPart"
    b64 = base64.b64encode(attachment_bytes).decode("utf-8")

    parts = [
        f"From: {sender}",
        f"To: {', '.join(recipients)}",
        f"Subject: {subject}",
        "MIME-Version: 1.0",
        f'Content-Type: multipart/mixed; boundary="{boundary}"',
        "",
        f"--{boundary}",
        'Content-Type: text/html; charset="UTF-8"',
        "",
        html,
        f"--{boundary}",
        f'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; name="{filename}"',
        "Content-Transfer-Encoding: base64",
        f'Content-Disposition: attachment; filename="{filename}"',
        "",
    ]

    for i in range(0, len(b64), 76):
        parts.append(b64[i:i + 76])

    parts.append(f"--{boundary}--")

    return "\n".join(parts).encode("utf-8")


# ----------------------------------------------------------------------
# HELPERS
# ----------------------------------------------------------------------
def combine_targets(rule):
    parts = []

    for ip in rule.get("IpRanges", []):
        if ip.get("CidrIp"):
            parts.append(ip.get("CidrIp"))

    for ip6 in rule.get("Ipv6Ranges", []):
        if ip6.get("CidrIpv6"):
            parts.append(ip6.get("CidrIpv6"))

    for pair in rule.get("UserIdGroupPairs", []):
        if pair.get("GroupId"):
            parts.append(pair.get("GroupId"))

    seen = set()
    ordered = []

    for p in parts:
        if p not in seen:
            seen.add(p)
            ordered.append(p)

    return ", ".join(ordered)


# ----------------------------------------------------------------------
# MAIN LAMBDA HANDLER
# ----------------------------------------------------------------------
def lambda_handler(event, context):
    sg_ids = event.get("security_group_ids", [])
    if not sg_ids:
        return {"statusCode": 400,
                "body": json.dumps({"error": "security_group_ids is required in event JSON"})}

    # Describe SGs
    try:
        resp = ec2.describe_security_groups(GroupIds=sg_ids)
        sg_map = {sg["GroupId"]: sg for sg in resp.get("SecurityGroups", [])}
    except ClientError as e:
        logger.exception("Failed to describe security groups")
        return {"statusCode": 500, "body": json.dumps({"error": str(e)})}

    inbound_rows = []
    outbound_rows = []
    in_serial = 1
    out_serial = 1

    # ------------------------------------------------------------------
    # PROCESS SECURITY GROUPS IN ORDER
    # ------------------------------------------------------------------
    for sg_id in sg_ids:
        sg = sg_map.get(sg_id)
        if not sg:
            inbound_rows.append({"Separator": True})
            outbound_rows.append({"Separator": True})
            continue

        group_name = sg.get("GroupName", "")

        # ==========================
        # INBOUND RULES
        # ==========================
        ip_perms = sg.get("IpPermissions", [])
        if ip_perms:
            for rule in ip_perms:
                port_range = "All"
                if "FromPort" in rule and "ToPort" in rule:
                    if rule["FromPort"] == rule["ToPort"]:
                        port_range = str(rule["FromPort"])
                    else:
                        port_range = f"{rule['FromPort']}-{rule['ToPort']}"

                proto = rule.get("IpProtocol", "All")
                if proto == "-1":
                    proto = "All"

                targets = []

                for ip in rule.get("IpRanges", []):
                    targets.append((ip.get("CidrIp"), ip.get("Description", "")))

                for ip6 in rule.get("Ipv6Ranges", []):
                    targets.append((ip6.get("CidrIpv6"), ip6.get("Description", "")))

                for pair in rule.get("UserIdGroupPairs", []):
                    targets.append((pair.get("GroupId"), pair.get("Description", "")))

                if not targets:
                    targets = [("", "")]

                for tgt, desc in targets:
                    inbound_rows.append({
                        "SrNo": in_serial,
                        "GroupName": group_name,
                        "GroupId": sg_id,
                        "Type": "Inbound",
                        "PortRange": port_range,
                        "Protocol": proto,
                        "Target": tgt,
                        "Description": desc
                    })
                    in_serial += 1

        inbound_rows.append({"Separator": True})

        # ==========================
        # OUTBOUND RULES
        # ==========================
        ip_perms_out = sg.get("IpPermissionsEgress", [])
        if ip_perms_out:
            for rule in ip_perms_out:
                port_range = "All"
                if "FromPort" in rule and "ToPort" in rule:
                    if rule["FromPort"] == rule["ToPort"]:
                        port_range = str(rule["FromFromPort"])
                    else:
                        port_range = f"{rule['FromPort']}-{rule['ToPort']}"

                proto = rule.get("IpProtocol", "All")
                if proto == "-1":
                    proto = "All"

                targets = []

                for ip in rule.get("IpRanges", []):
                    targets.append((ip.get("CidrIp"), ip.get("Description", "")))

                for ip6 in rule.get("Ipv6Ranges", []):
                    targets.append((ip6.get("CidrIpv6"), ip6.get("Description", "")))

                for pair in rule.get("UserIdGroupPairs", []):
                    targets.append((pair.get("GroupId"), pair.get("Description", "")))

                if not targets:
                    targets = [("", "")]

                for tgt, desc in targets:
                    outbound_rows.append({
                        "SrNo": out_serial,
                        "GroupName": group_name,
                        "GroupId": sg_id,
                        "Type": "Outbound",
                        "PortRange": port_range,
                        "Protocol": proto,
                        "Target": tgt,
                        "Description": desc
                    })
                    out_serial += 1

        outbound_rows.append({"Separator": True})

    # ------------------------------------------------------------------
    # BUILD WORKBOOK
    # ------------------------------------------------------------------
    wb = build_workbook(inbound_rows, outbound_rows)
    tmpfile = f"/tmp/{ATTACHMENT_FILENAME}"

    try:
        wb.save(tmpfile)
    except Exception as e:
        logger.exception("Failed to save workbook")
        return {"statusCode": 500, "body": json.dumps({"error": str(e)})}

    with open(tmpfile, "rb") as f:
        xbytes = f.read()

    # Get AWS Account ID
    try:
        account_id = sts.get_caller_identity()["Account"]
    except Exception:
        account_id = "Unknown"

    # ------------------------------------------------------------------
    # HTML EMAIL (same styling as IAM email)
    # ------------------------------------------------------------------
    html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; margin:0; padding:20px; background:#f5f6fa;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr><td align="center">
          <table width="720" cellpadding="0" cellspacing="0" 
                 style="background:#ffffff; border-radius:8px; padding:20px; 
                 box-shadow:0 2px 8px rgba(0,0,0,0.08);">

            <tr>
              <td align="center" style="padding-bottom:12px;">
                <img src="{PHI_LOGO_URL}" alt="PhiCommerce" height="60" style="display:block;">
              </td>
            </tr>

            <tr>
              <td align="center" style="padding-bottom:16px;">
                <div style="font-size:20px; font-weight:700; color:#222;">
                  AWS Account ID: {account_id}
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding-bottom:18px;">
                <table width="100%" cellpadding="10" cellspacing="0"
                       style="border-collapse:collapse; font-size:15px;">
                  <tr style="background:#f0f0f0;">
                    <th align="left" style="padding:10px;">Summary</th>
                    <th align="right" style="padding:10px;">Count</th>
                  </tr>

                  <tr>
                    <td style="padding:10px;">Total Inbound Rules</td>
                    <td align="right" style="padding:10px;">
                      <b>{len([r for r in inbound_rows if not r.get('Separator')])}</b>
                    </td>
                  </tr>

                  <tr>
                    <td style="padding:10px;">Total Outbound Rules</td>
                    <td align="right" style="padding:10px;">
                      <b>{len([r for r in outbound_rows if not r.get('Separator')])}</b>
                    </td>
                  </tr>

                </table>
              </td>
            </tr>

            <tr>
              <td style="color:#555; font-size:14px;">
                Attached is the Security Group Rules report.
              </td>
            </tr>

          </table>
        </td></tr>
      </table>
    </body>
    </html>
    """

    # SEND EMAIL
    raw = build_raw_email(SENDER, RECIPIENTS, SUBJECT, html, xbytes, ATTACHMENT_FILENAME)

    try:
        resp = ses.send_raw_email(RawMessage={"Data": raw})
        logger.info(f"Email sent, MessageId: {resp.get('MessageId')}")
    except ClientError as e:
        logger.exception("Failed to send email via SES")
        return {"statusCode": 500, "body": json.dumps({"error": str(e)})}

    return {
        "statusCode": 200,
        "body": json.dumps({
            "message": "Security group rules report generated and emailed",
            "inbound_rules": len([r for r in inbound_rows if not r.get("Separator")]),
            "outbound_rules": len([r for r in outbound_rows if not r.get("Separator")])
        })
    }
