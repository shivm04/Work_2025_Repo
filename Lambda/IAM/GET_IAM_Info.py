import boto3
import json
import logging
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from botocore.exceptions import ClientError

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# AWS clients
iam = boto3.client('iam')
sts = boto3.client('sts')
ses = boto3.client('ses')

# CONFIG - edit if needed
SENDER = "04shivm@gmail.com"           # SES verified sender (if in sandbox)
RECIPIENTS = ["04shivm@gmail.com"]    # recipients (verify if in sandbox)
SUBJECT = "AWS IAM Users & Groups Report"
ATTACHMENT_FILENAME = "iam_users_groups_report.xlsx"

# Logos (user-provided PhiCommerce webp + recommended AWS PNG)
PHI_LOGO_URL = "https://test.com/wp-content/uploads/2024/09/logo-300x170.png.webp"
#AWS_LOGO_URL = "https://a0.awsstatic.com/main/images/logos/aws-logo-color.png"

# -------------------------
# Helper functions
# -------------------------
def user_console_access(username: str) -> str:
    try:
        iam.get_login_profile(UserName=username)
        return "Yes"
    except iam.exceptions.NoSuchEntityException:
        return "No"
    except ClientError:
        return "Unknown"

def user_mfa_status(username: str) -> str:
    try:
        resp = iam.list_mfa_devices(UserName=username)
        return "Enabled" if resp.get("MFADevices") else "Disabled"
    except ClientError:
        return "Unknown"

def list_groups_for_user(username: str) -> list:
    groups = []
    paginator = iam.get_paginator("list_groups_for_user")
    for page in paginator.paginate(UserName=username):
        groups.extend([g["GroupName"] for g in page.get("Groups", [])])
    return groups

def list_attached_group_policy_names(group_name: str) -> list:
    names = []
    paginator = iam.get_paginator("list_attached_group_policies")
    for page in paginator.paginate(GroupName=group_name):
        for p in page.get("AttachedPolicies", []):
            names.append(p.get("PolicyName"))
    return names

def list_all_groups() -> list:
    groups = []
    paginator = iam.get_paginator("list_groups")
    for page in paginator.paginate():
        groups.extend(page.get("Groups", []))
    return groups

def list_users_in_group(group_name: str) -> list:
    users = []
    paginator = iam.get_paginator("get_group")
    for page in paginator.paginate(GroupName=group_name):
        for u in page.get("Users", []):
            users.append(u.get("UserName"))
    return users

# Excel helpers: style and auto column width
def style_header_row(ws, header_row=1):
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
    thin = Side(border_style="thin", color="FFBBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.font = bold
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_table_borders(ws):
    thin = Side(border_style="thin", color="FFBBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            # left align data
            cell.alignment = Alignment(horizontal="left", vertical="center")

def autosize_columns(ws, min_width=8, max_width=50):
    # compute max length per column
    dims = {}
    for row in ws.rows:
        for cell in row:
            val = cell.value
            if val is None:
                l = 0
            else:
                l = len(str(val))
            col = cell.column
            dims[col] = max(dims.get(col, 0), l)
    for col, max_len in dims.items():
        width = max(min_width, min(max_len + 2, max_width))
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = width

# Build workbook with formatting
def build_workbook(user_rows: list, group_rows: list) -> Workbook:
    wb = Workbook()

    # ---------------------------
    # IAM USER REPORT (with Sr. No.)
    # ---------------------------
    ws1 = wb.active
    ws1.title = "IAM User Report"

    headers_user = ["Sr. No.", "User", "ARN", "Console Access", "MFA", "Groups"]
    ws1.append(headers_user)

    sr = 1
    for r in user_rows:
        ws1.append([
            sr,
            r.get("UserName", ""),
            r.get("Arn", ""),
            r.get("ConsoleAccess", ""),
            r.get("MFA", ""),
            r.get("Groups", ""),
        ])
        sr += 1

    style_header_row(ws1, header_row=1)
    apply_table_borders(ws1)
    autosize_columns(ws1)

    # ---------------------------
    # IAM GROUP REPORT (with Sr. No.)
    # ---------------------------
    ws2 = wb.create_sheet(title="IAM Group Report")

    headers_group = ["Sr. No.", "Group Name", "Users", "Attached Policies"]
    ws2.append(headers_group)

    sr = 1
    for g in group_rows:
        ws2.append([
            sr,
            g.get("GroupName", ""),
            g.get("Users", ""),
            g.get("AttachedPolicies", ""),
        ])
        sr += 1

    style_header_row(ws2, header_row=1)
    apply_table_borders(ws2)
    autosize_columns(ws2)

    return wb

# Create raw MIME email with XLSX attachment
def create_raw_email_with_attachment(sender: str, recipients: list, subject: str, html_body: str, attachment_bytes: bytes, filename: str) -> bytes:
    boundary = "NextPart"
    attachment_b64 = base64.b64encode(attachment_bytes).decode("utf-8")

    parts = []
    parts.append(f"From: {sender}")
    parts.append(f"To: {', '.join(recipients)}")
    parts.append(f"Subject: {subject}")
    parts.append("MIME-Version: 1.0")
    parts.append(f'Content-Type: multipart/mixed; boundary="{boundary}"')
    parts.append("")
    parts.append(f"--{boundary}")
    parts.append('Content-Type: text/html; charset="UTF-8"')
    parts.append("Content-Transfer-Encoding: 7bit")
    parts.append("")
    parts.append(html_body)
    parts.append(f"--{boundary}")
    parts.append(f'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; name="{filename}"')
    parts.append("Content-Transfer-Encoding: base64")
    parts.append(f'Content-Disposition: attachment; filename="{filename}"')
    parts.append("")
    # wrap the base64 into 76-char lines per RFC
    for i in range(0, len(attachment_b64), 76):
        parts.append(attachment_b64[i:i+76])
    parts.append(f"--{boundary}--")

    raw = "\n".join(parts).encode("utf-8")
    return raw

# -------------------------
# Lambda handler
# -------------------------
def lambda_handler(event, context):
    logger.info("Starting IAM users & groups report")

    # Get account ID
    try:
        aws_account_id = sts.get_caller_identity()["Account"]
    except Exception:
        aws_account_id = "Unknown"

    # 1) Users
    users = []
    paginator = iam.get_paginator("list_users")
    for page in paginator.paginate():
        for u in page.get("Users", []):
            users.append({"UserName": u.get("UserName"), "Arn": u.get("Arn", "")})

    user_infos = []
    for u in users:
        uname = u["UserName"]
        groups = list_groups_for_user(uname)
        groups_str = ", ".join(sorted(groups)) if groups else ""
        user_infos.append({
            "UserName": uname,
            "Arn": u.get("Arn", ""),
            "ConsoleAccess": user_console_access(uname),
            "MFA": user_mfa_status(uname),
            "Groups": groups_str
        })

    # Sort users by group (alphabetical). Users without groups last.
    def _group_key(item):
        g = item.get("Groups", "")
        return (0, g.lower()) if g else (1, item.get("UserName", "").lower())
    user_infos_sorted = sorted(user_infos, key=_group_key)

    # 2) Groups (all attached policies)
    groups = list_all_groups()
    group_rows = []
    for g in groups:
        gname = g.get("GroupName")
        members = list_users_in_group(gname)
        members_str = ", ".join(sorted(members)) if members else ""
        attached = list_attached_group_policy_names(gname)
        attached_str = ", ".join(sorted(attached)) if attached else ""
        group_rows.append({
            "GroupName": gname,
            "Users": members_str,
            "AttachedPolicies": attached_str
        })

    # 3) Build workbook
    wb = build_workbook(user_infos_sorted, group_rows)

    # Save workbook to /tmp and read bytes
    tmpfile = f"/tmp/{ATTACHMENT_FILENAME}"
    try:
        wb.save(tmpfile)
    except Exception as e:
        logger.exception(f"Failed to save workbook: {e}")
        return {"statusCode": 500, "body": "Failed to save workbook"}

    try:
        with open(tmpfile, "rb") as f:
            xbytes = f.read()
    except Exception as e:
        logger.exception(f"Failed to read generated workbook: {e}")
        return {"statusCode": 500, "body": "Failed to read workbook"}

    # -------------------------
    # HTML Email - logos via URLs, same height on white background
    # Account ID bold centered, summary table below
    # -------------------------
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; margin:0; padding:20px; background:#f5f6fa;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr><td align="center">

          <table width="720" cellpadding="0" cellspacing="0" style="background:#ffffff; border-radius:8px; padding:20px; box-shadow:0 2px 8px rgba(0,0,0,0.08);">
            <!-- logo row on white background -->
            <tr>
              <td align="center" style="padding-bottom:12px; background-color:#ffffff;">
                <!-- use table to ensure white background for logos -->
                <table cellpadding="0" cellspacing="0" style="background:#ffffff;">
                  <tr>
                    <td style="padding-right:18px; background:#ffffff;">
                      <img src="{PHI_LOGO_URL}" alt="PhiCommerce" height="60" style="display:block; background:#ffffff;">
                    </td>
                  </tr>
                </table>
              </td>
            </tr>

            <!-- Account ID big & bold -->
            <tr>
              <td align="center" style="padding-bottom:16px;">
                <div style="font-size:20px; font-weight:700; color:#222;">AWS Account ID: {aws_account_id}</div>
              </td>
            </tr>

            <!-- Summary table -->
            <tr>
              <td style="padding-bottom:18px;">
                <table width="100%" cellpadding="10" cellspacing="0" style="border-collapse:collapse; font-size:15px;">
                  <tr style="background:#f0f0f0;">
                    <th align="left" style="padding:10px; border-bottom:2px solid #ddd;">Summary</th>
                    <th align="right" style="padding:10px; border-bottom:2px solid #ddd;">Count</th>
                  </tr>
                  <tr>
                    <td style="padding:10px; border-bottom:1px solid #eee;">Total IAM Users</td>
                    <td align="right" style="padding:10px; border-bottom:1px solid #eee;"><b>{len(user_infos_sorted)}</b></td>
                  </tr>
                  <tr>
                    <td style="padding:10px; border-bottom:1px solid #eee;">Total IAM Groups</td>
                    <td align="right" style="padding:10px; border-bottom:1px solid #eee;"><b>{len(group_rows)}</b></td>
                  </tr>
                </table>
              </td>
            </tr>

            <!-- Description -->
            <tr>
              <td style="color:#555; font-size:14px; padding-bottom:16px;">
                Attached is the detailed IAM Users & Groups report (XLSX).<br><br>
                <strong>Note:</strong> All group policies are listed by name.
              </td>
            </tr>

            <!-- Footer -->
            <tr>
              <td style="font-size:13px; color:#777;">
                Regards,<br><strong>Cloud Automation</strong>
              </td>
            </tr>

          </table>

        </td></tr>
      </table>
    </body>
    </html>
    """

    # Build raw email & send
    raw = create_raw_email_with_attachment(SENDER, RECIPIENTS, SUBJECT, html_body, xbytes, ATTACHMENT_FILENAME)
    try:
        resp = ses.send_raw_email(RawMessage={'Data': raw})
        logger.info(f"Email sent, MessageId: {resp.get('MessageId')}")
    except ClientError as e:
        logger.exception(f"Failed to send email via SES: {e}")
        return {"statusCode": 500, "body": "Failed to send email"}

    return {
        "statusCode": 200,
        "body": json.dumps({
            "message": "Report generated and emailed",
            "users": len(user_infos_sorted),
            "groups": len(group_rows)
        })
    }
